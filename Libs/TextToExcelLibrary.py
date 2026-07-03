import os
import re
import chardet
from openpyxl import load_workbook

class TextToExcelLibrary:
    """Library สำหรับแปลงไฟล์ .txt ไปเขียนลงในไฟล์ Excel (.xlsx) ที่มีอยู่แล้วเท่านั้น"""

    def file_encoding(self, file_path: str, num_bytes: int = 10000) -> str:
        """ตรวจหา encoding ของไฟล์แบบอัตโนมัติ"""
        with open(file_path, 'rb') as f:
            rawdata = f.read(num_bytes)
            result = chardet.detect(rawdata)
            return result['encoding'] or 'utf-8'
        
    def txt_to_excel(self,
                 input_txt: str,
                 output_xlsx: str,
                 delimiter: str = '|',
                 sheet_name: str = None,
                 start_row: int = 1,
                 start_col: int = 1,
                 apply_date: str = None) -> None:

        if apply_date:
            # input_txt = self.apply_date_to_path(input_txt, apply_date)
            output_xlsx = self.apply_date_to_path(output_xlsx, apply_date)

        if not os.path.exists(output_xlsx):
            raise FileNotFoundError(f"Excel file not found: {output_xlsx}")

        wb = load_workbook(output_xlsx)

        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        else:
            ws = wb.active

        # ลบข้อมูลเฉพาะใน column ที่กำหนด จาก start_row ลงไปจนเจอค่าว่าง
        row = start_row
        while True:
            cell = ws.cell(row=row, column=start_col)
            if cell.value is None:
                break
            for col in range(start_col, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
            row += 1

        encoding = self.file_encoding(input_txt)
        with open(input_txt, 'r', encoding=encoding, errors='ignore') as f:
            for idx, line in enumerate(f):
                if not line.strip():
                    continue
                cols = [val.strip() for val in line.rstrip('\n').split(delimiter)]
                for col_idx, val in enumerate(cols):
                    cell = ws.cell(row=start_row + idx, column=start_col + col_idx, value=val)
                    cell.number_format = '@'

        wb.save(output_xlsx)

    def apply_date_to_path(self, path: str, new_date: str) -> str:
        """แทนวันที่ใน path ที่กำหนดด้วยวันที่ใหม่"""
        return re.sub(r"\d{8}", new_date, path)

    def txt_to_excel_single_column(self,
                                   input_txt: str,
                                   output_xlsx: str,
                                   sheet_name: str = None,
                                   start_row: int = 1,
                                   start_col: int = 1,
                                   apply_date: str = None) -> None:
        """อ่านไฟล์ .txt แล้ววางลง Excel เฉพาะคอลัมน์เดียว โดยไม่แยกด้วย delimiter"""
        if apply_date:
            input_txt = self.apply_date_to_path(input_txt, apply_date)
            output_xlsx = self.apply_date_to_path(output_xlsx, apply_date)

        if not os.path.exists(output_xlsx):
            raise FileNotFoundError(f"Excel file not found: {output_xlsx}")

        wb = load_workbook(output_xlsx)

        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        else:
            ws = wb.active

        row = start_row
        while True:
            cell = ws.cell(row=row, column=start_col)
            if cell.value is None:
                break
            cell.value = None
            row += 1

        # with open(input_txt, 'r', encoding='utf-8') as f:
        #     for idx, line in enumerate(f):
        #         if not line.strip():
        #             continue
        #         ws.cell(row=start_row + idx, column=start_col, value=line.rstrip('\n')).number_format = '@'

        # wb.save(output_xlsx)
        encoding = self.file_encoding(input_txt)
        with open(input_txt, 'r', encoding=encoding, errors='ignore') as f:
        # with open(input_txt, 'r', encoding=encoding) as f:
            for idx, line in enumerate(f):
                if not line.strip():
                    continue
                ws.cell(row=start_row + idx, column=start_col, value=line.rstrip('\n')).number_format = '@'

        wb.save(output_xlsx)