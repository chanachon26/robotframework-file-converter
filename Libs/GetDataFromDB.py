import pymysql
import openpyxl
import os
from sshtunnel import SSHTunnelForwarder

def export_data_to_excel(db_host, db_user, db_password, db_name, db_port, query, output_file):
    db_port = int(db_port)
    connection = pymysql.connect(host=db_host,
                                 user=db_user,
                                 password=db_password,
                                 database=db_name,
                                 port=db_port)
    cursor = connection.cursor()
    cursor.execute(query)
    results = cursor.fetchall()
    column_names = [desc[0] for desc in cursor.description]

    directory = os.path.dirname(output_file)
    if not os.path.exists(directory):
        os.makedirs(directory)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Query Result"

    for col_num, column_name in enumerate(column_names, 1):
        sheet.cell(row=1, column=col_num, value=column_name)

    for row_num, row in enumerate(results, 2):
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(output_file)
    cursor.close()
    connection.close()

    print(f"Data exported to {output_file}")

def export_data_to_excel_via_ssh(
        ssh_host,
        ssh_port,
        ssh_user,
        ssh_password,

        db_host,
        db_user,
        db_password,
        db_name,
        db_port,

        query,
        output_file
):
    ssh_port = int(ssh_port)
    db_port = int(db_port)

    with SSHTunnelForwarder(
        (ssh_host, ssh_port),
        ssh_username=ssh_user,
        ssh_password=ssh_password,
        remote_bind_address=(db_host, db_port)
    ) as tunnel:

        connection = pymysql.connect(
            host='127.0.0.1',
            user=db_user,
            password=db_password,
            database=db_name,
            port=tunnel.local_bind_port,
            charset='utf8mb4'
        )

        cursor = connection.cursor()
        cursor.execute(query)
        results = cursor.fetchall()
        column_names = [desc[0] for desc in cursor.description]

        directory = os.path.dirname(output_file)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Query Result"

        for col_num, column_name in enumerate(column_names, 1):
            sheet.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(results, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(output_file)

        cursor.close()
        connection.close()

        print(f"Data exported to {output_file} (via SSH)")
